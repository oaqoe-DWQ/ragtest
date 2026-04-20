"""
Ragas评估结果详细导出模块
将每个样本的ragas指标导出到shuju文件夹的Excel文件中
"""

import os
import pandas as pd
from datetime import datetime
from typing import Dict, Any, List, Optional


def export_ragas_detail_to_excel(
    ragas_results: Dict[str, Any],
    output_path: Optional[str] = None
) -> str:
    """
    将Ragas评估结果按样本导出到Excel文件
    
    Args:
        ragas_results: Ragas评估结果字典（包含 raw_results 和各项指标）
        output_path: 输出路径，默认保存到 shuju/{数据集名称}_{timestamp}.xlsx
        
    Returns:
        str: 导出文件的完整路径
    """
    # 创建shuju目录
    shuju_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "shuju")
    os.makedirs(shuju_dir, exist_ok=True)
    
    # 生成文件名：使用数据集名称 + 时间戳
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        # 从 ragas_results 中获取数据集文件名称
        dataset_file = ragas_results.get('dataset_file', 'ragas_detail')
        # 移除文件扩展名
        if dataset_file.endswith('.xlsx'):
            dataset_name = dataset_file[:-5]
        elif dataset_file.endswith('.xls'):
            dataset_name = dataset_file[:-4]
        else:
            dataset_name = dataset_file if dataset_file else 'ragas_detail'
        output_path = os.path.join(shuju_dir, f"{dataset_name}_{timestamp}.xlsx")
    
    # 提取每个样本的分数
    sample_data = _extract_sample_scores(ragas_results)
    
    if not sample_data:
        # 如果没有提取到样本数据，创建一个空的或默认数据的文件
        sample_data = [{
            "样本ID": 1,
            "user_input": "",
            "response": "",
            "reference": "",
            "faithfulness": "",
            "answer_relevancy": "",
            "context_precision": "",
            "context_recall": "",
            "context_entity_recall": "",
            "context_relevance": "",
            "answer_correctness": "",
            "answer_similarity": "",
            "备注": "无有效评估数据"
        }]
    
    # 创建DataFrame
    df = pd.DataFrame(sample_data)
    
    # 调整列顺序
    columns_order = [
        "样本ID", "user_input", "response", "reference",
        "faithfulness", "answer_relevancy", "context_precision", "context_recall",
        "context_entity_recall", "context_relevance", "answer_correctness",
        "answer_similarity", "备注"
    ]
    
    # 只保留存在的列
    existing_columns = [col for col in columns_order if col in df.columns]
    df = df[existing_columns]
    
    # 保存到Excel（带格式）
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='样本详情', index=False)
        
        # 获取工作表
        worksheet = writer.sheets['样本详情']
        
        # 设置列宽
        column_widths = {
            'A': 10,   # 样本ID
            'B': 50,   # user_input
            'C': 50,   # response
            'D': 50,   # reference
            'E': 15,   # faithfulness
            'F': 18,   # answer_relevancy
            'G': 18,   # context_precision
            'H': 15,   # context_recall
            'I': 20,   # context_entity_recall
            'J': 17,   # context_relevance
            'K': 17,   # answer_correctness
            'L': 17,   # answer_similarity
            'M': 30,   # 备注
        }
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # 设置表头样式（加粗）
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.font = Font(bold=True, color="FFFFFF")
            cell.border = thin_border
        
        # 数据行样式
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(df) + 1), start=2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical='top', wrap_text=False)
                # 指标列居中对齐
                if cell.column_letter in ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    # 如果是数值，格式化
                    if isinstance(cell.value, (int, float)) and cell.value != "":
                        cell.number_format = '0.0000'
    
    return output_path


def _extract_sample_scores(ragas_results: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    从Ragas结果中提取每个样本的详细分数

    Args:
        ragas_results: 包含 raw_results 的字典

    Returns:
        List[Dict[str, Any]]: 每个样本的分数列表
    """
    sample_data = []
    raw_results = ragas_results.get('raw_results')

    # 获取原始样本数据（包含user_input、response、reference）
    original_samples = ragas_results.get('sample_data', [])

    if raw_results is None and not original_samples:
        return sample_data

    # 尝试从不同的数据结构中提取样本分数
    scores_df = None

    # 方式1: 如果是 EvaluationResult 对象本身
    if hasattr(raw_results, '__dict__'):
        obj_dict = raw_results.__dict__

        # 尝试获取 scores 属性
        if 'scores' in obj_dict:
            scores_data = obj_dict['scores']
            if hasattr(scores_data, 'to_pandas'):
                scores_df = scores_data.to_pandas()
            elif hasattr(scores_data, 'iloc'):
                scores_df = scores_data
            elif isinstance(scores_data, list):
                scores_df = pd.DataFrame(scores_data)

        # 尝试获取 dataset 属性
        if scores_df is None and 'dataset' in obj_dict:
            dataset = obj_dict['dataset']
            if hasattr(dataset, 'to_pandas'):
                scores_df = dataset.to_pandas()

    # 方式2: 如果 raw_results 本身就是字典
    if isinstance(raw_results, dict) and scores_df is None:
        if 'scores' in raw_results:
            scores_data = raw_results['scores']
            if hasattr(scores_data, 'to_pandas'):
                scores_df = scores_data.to_pandas()
            elif hasattr(scores_data, 'iloc'):
                scores_df = scores_data
            elif isinstance(scores_data, list):
                scores_df = pd.DataFrame(scores_data)

    # 方式3: 尝试 raw_results 是否有 to_pandas 方法
    if scores_df is None and hasattr(raw_results, 'to_pandas'):
        try:
            scores_df = raw_results.to_pandas()
        except Exception:
            pass

    metrics = [
        'faithfulness', 'answer_relevancy', 'context_precision', 'context_recall',
        'context_entity_recall', 'context_relevance', 'answer_correctness', 'answer_similarity'
    ]

    # 指标名别名映射：导出名 -> DataFrame中可能的列名列表
    # ChineseContextRelevance 的 name 为 "nv_context_relevance"，需要映射回来
    metric_column_variants = {
        'context_relevance': ['context_relevance', 'nv_context_relevance', 'Context Relevance'],
    }

    def _get_result_value(key: str) -> Any:
        """从 ragas_results 中获取值，支持列名别名"""
        if key in metric_column_variants:
            for variant in metric_column_variants[key]:
                val = ragas_results.get(variant)
                if val is not None:
                    return val
            return None
        return ragas_results.get(key)

    def _get_df_value(row: Any, metric_name: str) -> Any:
        """从 DataFrame 行中获取指标值，支持列名别名"""
        if not hasattr(row, 'get') and not hasattr(row, '__getitem__'):
            return None
        variants = metric_column_variants.get(metric_name, [metric_name])
        for col_name in variants:
            if hasattr(row, 'get'):
                value = row.get(col_name)
            else:
                value = None
            if value is not None:
                return value
        return None

    # 如果找到了 scores DataFrame，提取每个样本的分数
    if scores_df is not None and hasattr(scores_df, 'iterrows'):
        for idx, row in scores_df.iterrows():
            # 优先使用原始样本数据
            if idx < len(original_samples):
                orig = original_samples[idx]
                user_input = orig.get("user_input", "")
                response = orig.get("response", "")
                reference = orig.get("reference", "")
            else:
                user_input = ""
                response = ""
                reference = ""

            sample = {
                "样本ID": idx + 1,
                "user_input": user_input,
                "response": response,
                "reference": reference,
            }

            for metric in metrics:
                value = _get_df_value(row, metric)
                if value is not None and not (isinstance(value, float) and pd.isna(value)):
                    try:
                        sample[metric] = float(value)
                    except (ValueError, TypeError):
                        sample[metric] = ""
                else:
                    sample[metric] = ""

            sample["备注"] = ""
            sample_data.append(sample)

    # 如果没有从raw_results提取到分数，但有原始样本数据，按样本展开
    if not sample_data and original_samples:
        overall_scores = {m: _safe_float(_get_result_value(m)) for m in metrics}

        for idx, orig in enumerate(original_samples):
            sample = {
                "样本ID": idx + 1,
                "user_input": orig.get("user_input", ""),
                "response": orig.get("response", ""),
                "reference": orig.get("reference", ""),
            }
            for metric in metrics:
                sample[metric] = overall_scores.get(metric, "")
            sample["备注"] = "整体评估结果"
            sample_data.append(sample)

    # 如果既没有分数也没有样本数据，创建占位行
    if not sample_data:
        has_any_metric = any(_get_result_value(m) is not None for m in metrics)
        if has_any_metric:
            sample_data.append({
                "样本ID": 1,
                "user_input": "",
                "response": "",
                "reference": "",
                "faithfulness": _safe_float(_get_result_value('faithfulness')),
                "answer_relevancy": _safe_float(_get_result_value('answer_relevancy')),
                "context_precision": _safe_float(_get_result_value('context_precision')),
                "context_recall": _safe_float(_get_result_value('context_recall')),
                "context_entity_recall": _safe_float(_get_result_value('context_entity_recall')),
                "context_relevance": _safe_float(_get_result_value('context_relevance')),
                "answer_correctness": _safe_float(_get_result_value('answer_correctness')),
                "answer_similarity": _safe_float(_get_result_value('answer_similarity')),
                "备注": "整体评估结果，无样本级详情"
            })

    return sample_data


def _safe_float(value: Any) -> Any:
    """安全转换为浮点数，失败返回空字符串"""
    if value is None:
        return ""
    try:
        return float(value)
    except (ValueError, TypeError):
        return ""


def _get_metric_value(row_data: Any, metric_name: str) -> Any:
    """
    从数据行中获取指标值（支持多种可能的列名）
    
    Args:
        row_data: DataFrame行或其他支持get方法的对象
        metric_name: 指标名称
        
    Returns:
        指标值，如果未找到则返回None
    """
    if not hasattr(row_data, 'get'):
        return None
    
    # 指标可能的列名映射
    metric_variants = {
        'context_relevance': ['context_relevance', 'nv_context_relevance', 'Context Relevance'],
    }
    
    # 如果指标有多个可能的名称，依次尝试
    if metric_name in metric_variants:
        variants = metric_variants[metric_name]
    else:
        variants = [metric_name]
    
    for variant in variants:
        value = row_data.get(variant)
        if value is not None:
            return value
    
    return None


def get_latest_export_file() -> Optional[str]:
    """获取最新的导出文件路径"""
    shuju_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "shuju")
    if not os.path.exists(shuju_dir):
        return None
    
    files = [
        os.path.join(shuju_dir, f)
        for f in os.listdir(shuju_dir)
        if f.endswith('.xlsx') and not f.startswith('~$')
    ]
    
    if not files:
        return None
    
    return max(files, key=os.path.getmtime)


def list_export_files() -> List[Dict[str, Any]]:
    """列出所有导出文件"""
    shuju_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "shuju")
    if not os.path.exists(shuju_dir):
        return []
    
    files = []
    for f in os.listdir(shuju_dir):
        if f.endswith('.xlsx') and not f.startswith('~$'):
            filepath = os.path.join(shuju_dir, f)
            stat = os.stat(filepath)
            files.append({
                "filename": f,
                "path": filepath,
                "size": stat.st_size,
                "modified": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S")
            })
    
    return sorted(files, key=lambda x: x["modified"], reverse=True)


def export_overall_testreport(
    ragas_results: Dict[str, Any],
    output_path: Optional[str] = None
) -> str:
    """
    导出总体测试报告到 Overall_testreport 文件夹
    
    Args:
        ragas_results: Ragas评估结果字典（包含各项指标）
        output_path: 输出路径，默认保存到 Overall_testreport/{表格名称}_{timestamp}.xlsx
        
    Returns:
        str: 导出文件的完整路径
    """
    # 创建 Overall_testreport 目录
    overall_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Overall_testreport")
    os.makedirs(overall_dir, exist_ok=True)
    
    # 获取表格名称（从 dataset_file 字段）
    dataset_file = ragas_results.get('dataset_file', 'unknown')
    # 移除文件扩展名
    if dataset_file.endswith('.xlsx'):
        dataset_name = dataset_file[:-5]
    elif dataset_file.endswith('.xls'):
        dataset_name = dataset_file[:-4]
    else:
        dataset_name = dataset_file
    
    # 生成文件名：表格名称_时间戳
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(overall_dir, f"{dataset_name}_{timestamp}.xlsx")
    
    # 指标定义（英文名、中文名、指标key）
    metrics_info = [
        ("Faithfulness", "忠实度", "faithfulness"),
        ("Answer Relevancy", "回答相关性", "answer_relevancy"),
        ("Context Precision", "上下文精确度", "context_precision"),
        ("Context Recall", "上下文召回率", "context_recall"),
        ("Context Entity Recall", "上下文实体召回率", "context_entity_recall"),
        ("Context Relevance", "上下文相关性", "context_relevance"),
        ("Answer Correctness", "回答正确性", "answer_correctness"),
        ("Answer Similarity", "回答相似度", "answer_similarity"),
    ]
    
    # 指标名别名映射：导出名 -> ragas_results中可能的键名
    # ChineseContextRelevance 的 name 为 "nv_context_relevance"，需要映射回来
    overall_metric_variants = {
        'context_relevance': ['context_relevance', 'nv_context_relevance'],
    }

    # 提取指标值
    report_data = []
    valid_metrics = []

    for eng_name, cn_name, metric_key in metrics_info:
        # 支持别名查找
        keys_to_try = overall_metric_variants.get(metric_key, [metric_key])
        value = None
        for k in keys_to_try:
            v = ragas_results.get(k)
            if v is not None:
                value = v
                break
        if value is not None:
            try:
                score = float(value)
                percentage = f"{score * 100:.1f}%"
                status = "正常" if not ragas_results.get('fallback_mode', False) else "Fallback"
            except (ValueError, TypeError):
                score = "N/A"
                percentage = "N/A"
                status = "评估失败"
            valid_metrics.append((eng_name, cn_name, score, percentage, status))
        else:
            valid_metrics.append((eng_name, cn_name, "N/A", "N/A", "评估失败"))
    
    # 计算统计数据
    valid_scores = [item[2] for item in valid_metrics if isinstance(item[2], float)]
    if valid_scores:
        avg_score = sum(valid_scores) / len(valid_scores)
        avg_percentage = f"{avg_score * 100:.1f}%"
        max_metric = max(valid_metrics, key=lambda x: x[2] if isinstance(x[2], float) else 0)
        min_metric = min(valid_metrics, key=lambda x: x[2] if isinstance(x[2], float) else 1)
        valid_count = len(valid_scores)
    else:
        avg_score = "N/A"
        avg_percentage = "N/A"
        max_metric = ("N/A", "N/A", "N/A")
        min_metric = ("N/A", "N/A", "N/A")
        valid_count = 0
    
    total_metrics = len(metrics_info)
    fallback_mode = ragas_results.get('fallback_mode', False)
    error_message = ragas_results.get('error_message', '')
    
    # 构建报告内容
    report_rows = []
    
    # 表头
    report_rows.append(["=" * 60])
    report_rows.append(["Ragas 评估总体测试报告"])
    report_rows.append(["=" * 60])
    report_rows.append([])
    
    # 基本信息
    report_rows.append(["基本信息"])
    report_rows.append(["-" * 40])
    report_rows.append([f"数据集文件: {dataset_file}"])
    report_rows.append([f"评估时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    report_rows.append([f"评估模式: {'Fallback模式' if fallback_mode else '正常模式'}"])
    if error_message:
        report_rows.append([f"错误信息: {error_message}"])
    report_rows.append([])
    
    # 指标汇总表
    report_rows.append(["指标汇总"])
    report_rows.append(["-" * 60])
    report_rows.append([f"{'指标名称':<25} {'中文名称':<15} {'分数':<10} {'百分比':<10} {'状态':<10}"])
    report_rows.append(["-" * 60])
    
    for eng_name, cn_name, score, percentage, status in valid_metrics:
        score_str = f"{score:.4f}" if isinstance(score, float) else str(score)
        report_rows.append([f"{eng_name:<25} {cn_name:<15} {score_str:<10} {percentage:<10} {status:<10}"])
    
    report_rows.append(["-" * 60])
    report_rows.append([])
    
    # 详细分析
    report_rows.append(["详细分析"])
    report_rows.append(["-" * 40])
    report_rows.append([f"• 平均分数: {avg_score:.4f} ({avg_percentage})" if isinstance(avg_score, float) else f"• 平均分数: {avg_score}"])
    report_rows.append([f"• 有效指标数: {valid_count}/{total_metrics}"])
    
    if valid_scores and not fallback_mode:
        report_rows.append([f"• 最高分数: {max_metric[1]} ({max_metric[2]:.4f})"])
        report_rows.append([f"• 最低分数: {min_metric[1]} ({min_metric[2]:.4f})"])
    elif fallback_mode:
        report_rows.append(["• 警告: 当前为fallback模式，分数为默认值"])
    
    report_rows.append([])
    report_rows.append(["=" * 60])
    
    # 创建DataFrame用于Excel导出
    df = pd.DataFrame({
        '指标名称': [item[0] for item in valid_metrics],
        '中文名称': [item[1] for item in valid_metrics],
        '分数': [f"{item[2]:.4f}" if isinstance(item[2], float) else item[2] for item in valid_metrics],
        '百分比': [item[3] for item in valid_metrics],
        '状态': [item[4] for item in valid_metrics]
    })
    
    # 保存到Excel（带格式）
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Sheet1: 指标汇总
        df.to_excel(writer, sheet_name='指标汇总', index=False)
        
        # 获取工作表
        worksheet = writer.sheets['指标汇总']
        
        # 设置列宽
        column_widths = {'A': 25, 'B': 15, 'C': 12, 'D': 12, 'E': 15}
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # 设置表头样式
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 数据行样式
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(df) + 1), start=2):
            for cell in row:
                cell.border = thin_border
                if cell.column_letter in ['C', 'D']:
                    cell.alignment = Alignment(horizontal='center')
    
    return output_path


def get_overall_report_files() -> List[Dict[str, Any]]:
    """列出所有总体测试报告文件"""
    overall_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Overall_testreport")
    if not os.path.exists(overall_dir):
        return []
    
    files = []
    for f in os.listdir(overall_dir):
        if f.endswith('.xlsx'):
            filepath = os.path.join(overall_dir, f)
            stat = os.stat(filepath)
            files.append({
                "filename": f,
                "path": filepath,
                "size": stat.st_size,
                "modified": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S")
            })
    
    return sorted(files, key=lambda x: x["modified"], reverse=True)
